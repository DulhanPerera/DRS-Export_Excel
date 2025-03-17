import configparser
from datetime import datetime, timedelta
import logging
import logging.config
import os
import sys
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from bson import ObjectId

# Load logger configuration
logging.config.fileConfig('Config/logger/loggers.ini')
logger = logging.getLogger('excel_data_writer')

# Load styles globally
config = configparser.ConfigParser()
try:
    config.read(os.path.join(os.path.dirname(__file__), '../Config/styles.ini'))
    if not config.has_section("MAIN_HEADER_FILLS"):
        logger.error("Section 'MAIN_HEADER_FILLS' not found in styles.ini.")
        sys.exit(1)
except Exception as e:
    logger.error(f"Failed to load styles.ini: {e}")
    sys.exit(1)

#region Styles for tables and fonts
styles = {
    "header_font": Font(
        bold=config.getboolean("FONTS", "header_font_bold"),
        color=config.get("FONTS", "header_font_color"),
        size=config.getint("FONTS", "header_font_size")
    ),
    "bold_font": Font(
        bold=config.getboolean("FONTS", "bold_font_bold")
    ),
    "main_header_fill": PatternFill(
        start_color=config.get("MAIN_HEADER_FILLS", "main_header_fill_start_color"),
        end_color=config.get("MAIN_HEADER_FILLS", "main_header_fill_end_color"),
        fill_type=config.get("MAIN_HEADER_FILLS", "main_header_fill_type")
    ),
    "sub_header_fill": PatternFill(
        start_color=config.get("SUB_HEADER_FILLS", "header_fill_start_color"),
        end_color=config.get("SUB_HEADER_FILLS", "header_fill_end_color"),
        fill_type=config.get("SUB_HEADER_FILLS", "header_fill_type")
    ),
    "cell_border": Border(
        left=Side(style=config.get("BORDERS", "cell_border_left_style")),
        right=Side(style=config.get("BORDERS", "cell_border_right_style")),
        top=Side(style=config.get("BORDERS", "cell_border_top_style")),
        bottom=Side(style=config.get("BORDERS", "cell_border_bottom_style"))
    ),
    "sub_header_alignment": Alignment(
        horizontal=config.get("SUB_HEADER_ALIGNMENTS", "header_alignment_horizontal"),
        vertical=config.get("SUB_HEADER_ALIGNMENTS", "header_alignment_vertical")
    ),
    "main_header_alignment": Alignment(
        horizontal=config.get("MAIN_HEADER_ALIGNMENTS", "main_header_alignment_horizontal"),
        vertical=config.get("MAIN_HEADER_ALIGNMENTS", "main_header_alignment_vertical")
    )
}
#endregion

def get_config():
    """
    Load and return the configuration from the Config.ini file.
    """
    try:
        config = configparser.ConfigParser()
        config.read(os.path.join(os.path.dirname(__file__), '../Config/Config.ini'))
        if not config.sections():
            logger.error("Configuration file is empty or not found.")
            sys.exit(1)
        return config
    except Exception as e:
        logger.error(f"Failed to load configuration: {e}")
        sys.exit(1)

def connect_db(config):
    """
    Connect to the MongoDB database using the configuration.
    """
    try:
        client = MongoClient(config['DATABASE']['MONGO_URI'])
        db = client[config['DATABASE']['DB_NAME']]
        logger.info("Successfully connected to the database.")
        return db
    except Exception as e:
        logger.error(f"Failed to connect to the database: {e}")
        sys.exit(1)

def format_with_thousand_separator(value):
    """
    Format numeric values with thousand separators.
    """
    if isinstance(value, (int, float)):
        return f"{value:,}"  # Add thousand separators
    return value

def get_arrears_band_value(db, current_arrears_band):
    """
    Retrieve the value for the given arrears band from the arrears_bands collection.
    """
    try:
        arrears_bands_collection = db["Arrears_bands"]
        arrears_bands_doc = arrears_bands_collection.find_one({})
        if arrears_bands_doc:
            return arrears_bands_doc.get(current_arrears_band)
        else:
            logger.warning("No arrears bands document found in the collection.")
            return None
    except Exception as e:
        logger.error(f"Failed to retrieve arrears band value: {e}")
        return None

def get_settlement_data(db, case_id):
    """
    Retrieve settlement data for the given case_id from the case_settlements collection.
    """
    try:
        settlements_collection = db["Case_settlements"]
        settlements = list(settlements_collection.find({"case_id": case_id}))
        if settlements:
            logger.info(f"Found {len(settlements)} settlement records for case_id: {case_id}")
        else:
            logger.warning(f"No settlement records found for case_id: {case_id}")
        return settlements
    except Exception as e:
        logger.error(f"Failed to retrieve settlement data: {e}")
        return []

def get_settlement_plan_data(db, case_id):
    """
    Retrieve settlement plan data for the given case_id from the case_settlements collection.
    """
    try:
        settlements_collection = db["Case_settlements"]
        settlements = list(settlements_collection.find({"case_id": case_id}))
        settlement_plans = []
        for settlement in settlements:
            if "settlement_plan" in settlement:
                for plan in settlement["settlement_plan"]:
                    # Add settlement_id to each plan
                    plan["settlement_id"] = settlement.get("settlement_id")
                    settlement_plans.append(plan)
        if settlement_plans:
            logger.info(f"Found {len(settlement_plans)} settlement plan records for case_id: {case_id}")
        else:
            logger.warning(f"No settlement plan records found for case_id: {case_id}")
        return settlement_plans
    except Exception as e:
        logger.error(f"Failed to retrieve settlement plan data: {e}")
        return []

def create_table(ws, start_row, start_col, main_header, sub_headers, data, styles):
    """
    Create a table with a start_process header, sub-headers, and data.
    """
    try:
        # Merge cells for the start_process header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(sub_headers) - 1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value=main_header)
        main_header_cell.font = styles["header_font"]
        main_header_cell.fill = styles["main_header_fill"]
        main_header_cell.border = styles["cell_border"]
        main_header_cell.alignment = styles["main_header_alignment"]
        
        # Write sub-headers
        for idx, header in enumerate(sub_headers, start=0):
            sub_header_cell = ws.cell(row=start_row + 1, column=start_col + idx, value=header)
            sub_header_cell.font = styles["header_font"]
            sub_header_cell.fill = styles["sub_header_fill"]
            sub_header_cell.border = styles["cell_border"]
            sub_header_cell.alignment = styles["sub_header_alignment"]
        
        # Insert data
        for data_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=0):
                data_cell = ws.cell(row=start_row + 1 + data_idx, column=start_col + col_idx, value=value)
                data_cell.border = styles["cell_border"]
        
        # Adjust column widths
        for col in range(start_col, start_col + len(sub_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Table '{main_header}' created successfully.")
        return start_row + len(data) + 3
    except Exception as e:
        logger.error(f"Failed to create table: {e}")
        sys.exit(1)

def create_case_details_table(ws, case_data, start_row, start_col, db):
    """
    Create the Case Details table in the worksheet with a horizontal layout and a start_process header.
    """
    try:
        logger.info("Creating Case Details table...")
        
        # Define the start_process header
        main_header = "Case Details"
        
        # Write the start_process header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        main_header_cell = ws.cell(row=start_row, column=start_col, value=main_header)
        main_header_cell.font = styles["header_font"]
        main_header_cell.fill = styles["main_header_fill"]
        main_header_cell.border = styles["cell_border"]
        main_header_cell.alignment = styles["main_header_alignment"]
        
        # Move to the next row for the table
        start_row += 1
        
        # Define headers for the table
        headers = [
            "Case ID", "Incident ID", "Account No.", "Customer Ref", "Area",
            "BSS Arrears Amount", "Current Arrears Amount", "Action type", "Filtered reason",
            "Last Payment Date", "Last BSS Reading Date", "Commission", "Case Current Status",
            "Current Arrears band", "DRC Commission Rule", "Created dtm", "Implemented dtm",
            "RTOM", "Monitor months"
        ]
        
        # Map MongoDB data to headers
        data_mapping = {
            "Case ID": case_data.get("case_id"),
            "Incident ID": case_data.get("incident_id"),
            "Account No.": case_data.get("account_no"),
            "Customer Ref": case_data.get("customer_ref"),
            "Area": case_data.get("area"),
            "BSS Arrears Amount": format_with_thousand_separator(case_data.get("bss_arrears_amount")),
            "Current Arrears Amount": format_with_thousand_separator(case_data.get("current_arrears_amount")),
            "Action type": case_data.get("action_type"),
            "Filtered reason": case_data.get("filtered_reason"),
            "Last Payment Date": case_data.get("last_payment_date"),
            "Last BSS Reading Date": case_data.get("last_bss_reading_date"),
            "Commission": format_with_thousand_separator(case_data.get("commission")),
            "Case Current Status": case_data.get("case_current_status"),
            "Current Arrears band": case_data.get("current_arrears_band"),
            "DRC Commission Rule": case_data.get("drc_commision_rule"),
            "Created dtm": case_data.get("created_dtm"),
            "Implemented dtm": case_data.get("implemented_dtm"),
            "RTOM": case_data.get("rtom"),
            "Monitor months": case_data.get("monitor_months")
        }
        
        # Retrieve arrears band value
        current_arrears_band = case_data.get("current_arrears_band")
        if current_arrears_band:
            arrears_band_value = get_arrears_band_value(db, current_arrears_band)
            if arrears_band_value:
                data_mapping["Current Arrears band"] = arrears_band_value
            else:
                logger.warning(f"No value found for arrears band: {current_arrears_band}")
        
        # Write headers and data horizontally
        for idx, header in enumerate(headers):
            # Write header in the first column
            ws.cell(row=start_row + idx, column=start_col, value=header).font = styles["header_font"]
            ws.cell(row=start_row + idx, column=start_col).fill = styles["sub_header_fill"]
            ws.cell(row=start_row + idx, column=start_col).border = styles["cell_border"]
            ws.cell(row=start_row + idx, column=start_col).alignment = styles["sub_header_alignment"]
            
            # Write corresponding data in the next column
            value = data_mapping.get(header)
            if isinstance(value, (list, dict)):
                value = str(value)
            ws.cell(row=start_row + idx, column=start_col + 1, value=value).border = styles["cell_border"]
            if header in ["Case ID", "Incident ID"]:
                ws.cell(row=start_row + idx, column=start_col + 1).font = styles["bold_font"]
        
        # Adjust column widths
        for col in range(start_col, start_col + 2):  # Only two columns (headers and data)
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(start_row, start_row + len(headers)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Case Details table created successfully.")
        return start_row + len(headers) + 1
    except Exception as e:
        logger.error(f"Failed to create Case Details table: {e}")
        sys.exit(1)

def create_contact_details_table(ws, case_data, start_row, start_col):
    """
    Create the Contact Details table in the worksheet.
    """
    try:
        logger.info("Creating Contact Details table...")
        contacts_headers = ["Mobile", "Email", "Home Phone", "Address"]
        
        # Prepare data for the table
        contacts = case_data.get("contact", [])
        data = [[contact.get("mob"), contact.get("email"), contact.get("lan"), contact.get("address")] for contact in contacts]
        
        # Create the table
        next_row = create_table(ws, start_row, start_col, "Contact Info", contacts_headers, data, styles)
        
        logger.info("Contact Details table created successfully.")
        return next_row
    except Exception as e:
        logger.error(f"Failed to create Contact Details table: {e}")
        sys.exit(1)

def create_remarks_table(ws, case_data, start_row, start_col):
    """
    Create the Remarks table in the worksheet.
    """
    try:
        logger.info("Creating Remarks table...")
        headers = ["Remark", "Remark Added by", "Remark Added Date"]
        
        # Prepare data for the table
        remarks = case_data.get("remark", [])
        data = [[remark.get("remark"), remark.get("remark_added_by"), remark.get("remark_added_date")] for remark in remarks]
        
        # Create the table
        next_row = create_table(ws, start_row, start_col, "Remarks", headers, data, styles)
        
        logger.info("Remarks table created successfully.")
        return next_row
    except Exception as e:
        logger.error(f"Failed to create Remarks table: {e}")
        sys.exit(1)

def create_settlement_table(ws, settlements, start_row, start_col):
    """
    Create the Settlement table in the worksheet.
    """
    try:
        logger.info("Creating Settlement table...")
        headers = [
            "Settlement ID", "Case ID", "DRC Name", "RO Name", "Status", "Status reason",
            "Status DTM", "Settlement Type", "Settlement Amount", "Settlement Phase",
            "Settlement Created by", "Settlement Created DTM", "Last Monitoring DTM", "Remark"
        ]
        
        # Prepare data for the table
        data = [
            [
                settlement.get("settlement_id"), settlement.get("case_id"), settlement.get("drc_id"),
                settlement.get("ro_id"), settlement.get("settlement_status"), settlement.get("status_reason"),
                settlement.get("status_dtm"), settlement.get("settlement_type"), settlement.get("settlement_amount"),
                settlement.get("settlement_phase"), settlement.get("created_by"), settlement.get("created_on"),
                settlement.get("last_monitoring_dtm"), settlement.get("remark")
            ]
            for settlement in settlements
        ]
        
        # Create the table
        next_row = create_table(ws, start_row, start_col, "Settlement Details", headers, data, styles)
        
        logger.info("Settlement table created successfully.")
        return next_row
    except Exception as e:
        logger.error(f"Failed to create Settlement table: {e}")
        sys.exit(1)

def create_settlement_plan_table(ws, settlement_plans, start_row, start_col):
    """
    Create the Settlement Plan table in the worksheet.
    """
    try:
        logger.info("Creating Settlement Plan table...")
        headers = [
            "Settlement ID", "Installment Sequence", "Installment Settle Amount",
            "Accumulated Amount", "Plan Date and Time"
        ]
        
        # Prepare data for the table
        data = [
            [
                plan.get("settlement_id"), plan.get("installment_seq"), plan.get("installment_settle_amount"),
                plan.get("accumulated_amount"), plan.get("plan_date")
            ]
            for plan in settlement_plans
        ]
        
        # Create the table
        next_row = create_table(ws, start_row, start_col, "Settlement Plan", headers, data, styles)
        
        logger.info("Settlement Plan table created successfully.")
        return next_row
    except Exception as e:
        logger.error(f"Failed to create Settlement Plan table: {e}")
        sys.exit(1)

def create_all_tables(wb, case_data, db):
    """
    Create the Case Details sheet with all tables.
    """
    try:
        logger.info("Creating Case Details sheet...")
        ws = wb.active
        ws.title = "Case Details"
        
        # Define starting row and column for the first table
        start_row, start_col = 1, 1
        
        # Create the Case Details table
        next_row = create_case_details_table(ws, case_data, start_row, start_col, db)
        
        # Add a two-row gap between the tables
        gap_row = next_row + 2
        
        # Create the Contact Details table
        next_row = create_contact_details_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Contact Info and Remarks tables
        gap_row = next_row + 1
        
        # Create the Remarks table
        next_row = create_remarks_table(ws, case_data, gap_row, start_col)
        
        # Add a two-row gap between the Remarks and Settlement tables
        gap_row = next_row + 1
        
        # Retrieve settlement data for the case
        case_id = case_data.get("case_id")
        settlements = get_settlement_data(db, case_id)
        
        # Create the Settlement table if settlement data exists
        if settlements:
            next_row = create_settlement_table(ws, settlements, gap_row, start_col)
            gap_row = next_row + 1  # Add a gap after the Settlement table
        
        # Retrieve settlement plan data for the case
        settlement_plans = get_settlement_plan_data(db, case_id)
        
        # Create the Settlement Plan table if settlement plan data exists
        if settlement_plans:
            create_settlement_plan_table(ws, settlement_plans, gap_row, start_col)
        
        logger.info("Case Details sheet created successfully.")
        return ws
    except Exception as e:
        logger.error(f"Failed to create Case Details sheet: {e}")
        sys.exit(1)

def export_all_tables(db, incident_id, output_path, collection_name):
    """
    Export case details to an Excel file with the filename containing Incident ID, Date, and Time.
    If a file with the same name exists, a number is appended to the filename to avoid overwriting.
    """
    try:
        logger.info(f"Exporting case details for Incident ID: {incident_id}")
        case_collection = db[collection_name]
        case_data = case_collection.find_one({"incident_id": incident_id})
        
        if not case_data:
            logger.error(f"No case details found for Incident ID: {incident_id}")
            sys.exit(1)
        
        logger.info(f"Case data found: {case_data}")
        
        wb = Workbook()
        ws = create_all_tables(wb, case_data, db)
        
        # Get the current date and time
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        
        # Create the base output file name with incident_id and current date/time
        file_name = f"case_details_{incident_id}_{current_time}.xlsx"
        
        # Ensure the output path ends with the correct file name
        if not output_path.endswith('.xlsx'):
            output_path = os.path.join(output_path, file_name)
        else:
            output_path = os.path.join(output_path, file_name)
        
        # Check if the file exists and append a number to make the filename unique
        if os.path.exists(output_path):
            base, extension = os.path.splitext(output_path)
            counter = 1
            while os.path.exists(f"{base}_{counter}{extension}"):
                counter += 1
            output_path = f"{base}_{counter}{extension}"
        
        # Create directories if they don't exist
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save the workbook
        try:
            wb.save(output_path)
            logger.info(f"Case details exported to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            sys.exit(1)
    except Exception as e:
        logger.error(f"Failed to export case details: {e}")
        sys.exit(1)

def start_process():
    """
    Main function to execute the case details export process.
    """
    try:
        logger.info("Starting case details export process...")
        config = get_config()
        db = connect_db(config)
        incident_id = 78910  # Replace with the actual incident ID you want to export
        export_path = config['EXCEL_EXPORT_FOLDER']['WIN_DB']  # Updated to use EXCEL_EXPORT_FOLDER
        collection_name = config['COLLECTIONS']['CASE_DETAIL_COLLECTION_NAME']
        
        export_all_tables(db, incident_id, export_path, collection_name)
        
        logger.info("Case details export process completed.")
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    start_process()