import logging
import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

logger = logging.getLogger('excel_data_writer')

def create_table(worksheet, x_pointer, y_pointer, main_header, sub_headers, data, styles):
    """
    Create a table with a main header, sub-headers, and data.
    """
    try:
        # Merge cells for the main header
        worksheet.merge_cells(start_row=x_pointer, start_column=y_pointer, end_row=x_pointer, end_column=y_pointer + len(sub_headers) - 1)
        main_header_cell = worksheet.cell(row=x_pointer, column=y_pointer, value=main_header)
        main_header_cell.font = styles["header_font"]
        main_header_cell.fill = styles["main_header_fill"]
        main_header_cell.border = styles["cell_border"]
        main_header_cell.alignment = styles["main_header_alignment"]
        
        # Write sub-headers
        for index, header in enumerate(sub_headers, start=0):
            sub_header_cell = worksheet.cell(row=x_pointer + 1, column=y_pointer + index, value=header)
            sub_header_cell.font = styles["header_font"]
            sub_header_cell.fill = styles["sub_header_fill"]
            sub_header_cell.border = styles["cell_border"]
            sub_header_cell.alignment = styles["sub_header_alignment"]
        
        # Insert data
        for data_index, row_data in enumerate(data, start=1):
            for col_index, value in enumerate(row_data, start=0):
                data_cell = worksheet.cell(row=x_pointer + 1 + data_index, column=y_pointer + col_index, value=value)
                data_cell.border = styles["cell_border"]
        
        # Adjust column widths
        for col in range(y_pointer, y_pointer + len(sub_headers)):
            max_length = 0
            column_letter = chr(64 + col)
            for row in range(x_pointer, x_pointer + len(data) + 2):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Table '{main_header}' created successfully.")
        return x_pointer + len(data) + 3
    except Exception as failed_table_creation:
        logger.error(f"Failed to create table: {failed_table_creation}")
        sys.exit(1)