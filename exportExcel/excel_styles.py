import sys
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment  # Import styling classes from openpyxl
import configparser  # Module for reading configuration files
import os
import logging  # Module for logging errors and debug information

# Initialize logger for this module
logger = logging.getLogger('excel_data_writer')


def load_styles(styles_config_path):
    """
    Load and return styles from the specified styles configuration file.

    Args:
        styles_config_path (str): The path to the styles.ini configuration file.

    Returns:
        dict: A dictionary containing various styles for Excel formatting.

    Outputs:
        - Logs an error and exits if the 'MAIN_HEADER_FILLS' section is missing.
        - Logs an error and exits if any exception occurs during loading.

    Exceptions:
        - Exits the program if the configuration file is invalid or missing required sections.
    """
    try:
        # Initialize the configuration parser and read the styles configuration file
        config = configparser.ConfigParser()
        config.read(styles_config_path)

        # Ensure that the required section exists in the configuration file
        if not config.has_section("MAIN_HEADER_FILLS"):
            logger.error("Section 'MAIN_HEADER_FILLS' not found in styles.ini.")
            sys.exit(1)  # Exit the program if the section is missing

        # Create a dictionary to store various styles from the configuration file
        styles = {
            "header_font": Font(
                bold=config.getboolean("FONTS", "header_font_bold"),
                color=config.get("FONTS", "header_font_color"),
                size=config.getint("FONTS", "header_font_size")
            ),
            "bold_font": Font(
                bold=config.getboolean("FONTS", "bold_font_bold")
            ),
            "main_header_fill": PatternFill(
                start_color=config.get("MAIN_HEADER_FILLS", "main_header_fill_start_color"),
                end_color=config.get("MAIN_HEADER_FILLS", "main_header_fill_end_color"),
                fill_type=config.get("MAIN_HEADER_FILLS", "main_header_fill_type")
            ),
            "sub_header_fill": PatternFill(
                start_color=config.get("SUB_HEADER_FILLS", "header_fill_start_color"),
                end_color=config.get("SUB_HEADER_FILLS", "header_fill_end_color"),
                fill_type=config.get("SUB_HEADER_FILLS", "header_fill_type")
            ),
            "cell_border": Border(
                left=Side(style=config.get("BORDERS", "cell_border_left_style")),
                right=Side(style=config.get("BORDERS", "cell_border_right_style")),
                top=Side(style=config.get("BORDERS", "cell_border_top_style")),
                bottom=Side(style=config.get("BORDERS", "cell_border_bottom_style"))
            ),
            "sub_header_alignment": Alignment(
                horizontal=config.get("SUB_HEADER_ALIGNMENTS", "header_alignment_horizontal"),
                vertical=config.get("SUB_HEADER_ALIGNMENTS", "header_alignment_vertical")
            ),
            "main_header_alignment": Alignment(
                horizontal=config.get("MAIN_HEADER_ALIGNMENTS", "main_header_alignment_horizontal"),
                vertical=config.get("MAIN_HEADER_ALIGNMENTS", "main_header_alignment_vertical")
            )
        }

        return styles  # Return the loaded styles as a dictionary

    except Exception as styles_loading_error:
        # Log an error message and exit if an exception occurs while loading styles
        logger.error(f"Failed to load styles: {styles_loading_error}")
        sys.exit(1)


def format_with_thousand_separator(value):
    """
    Format numeric values with thousand separators.

    Args:
        value (int, float, or any): The value to be formatted.

    Returns:
        str: The formatted string with thousand separators for numbers.
        Original value if it's not an integer or float.

    Example:
        format_with_thousand_separator(1000000)  # Returns '1,000,000'
        format_with_thousand_separator(1234.56)  # Returns '1,234.56'
        format_with_thousand_separator("test")   # Returns 'test'

    Notes:
        - Uses Python's built-in string formatting to add thousand separators.
    """
    if isinstance(value, (int, float)):
        return f"{value:,}"  # Format numbers with commas as thousand separators
    return value  # Return the original value if it's not numeric
